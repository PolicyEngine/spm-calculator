"""
Generate bundled metro data and static web assets for the SPM calculator.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
import requests


def _read_package_version() -> str:
    """Read spm-calculator's version from pyproject.toml.

    Keeps the web bundle's ``packageVersion`` field in lockstep with
    the Python package's release tag without requiring tomllib.
    """
    repo_root = Path(__file__).resolve().parent.parent
    pyproject = (repo_root / "pyproject.toml").read_text()
    match = re.search(r'^version\s*=\s*"([^"]+)"', pyproject, re.MULTILINE)
    if not match:
        raise RuntimeError("Could not find a version field in pyproject.toml")
    return match.group(1)


from spm_calculator.equivalence_scale import REFERENCE_RAW_SCALE
from spm_calculator.forecast import (
    CPI_PROJECTIONS,
    LATEST_PUBLISHED_YEAR,
    forecast_thresholds,
    get_thresholds,
)
from spm_calculator.forecast import (
    HISTORICAL_THRESHOLDS as _PACKAGE_THRESHOLDS,
)
from spm_calculator.nowcast import get_nowcast_years, nowcast_with_metadata

# The web bundle uses string year keys throughout.
HISTORICAL_THRESHOLDS = {
    str(year): dict(tenures) for year, tenures in _PACKAGE_THRESHOLDS.items()
}

TENURE_HOUSING_SHARES = {
    "renter": 0.443,
    "owner_with_mortgage": 0.434,
    "owner_without_mortgage": 0.323,
}

METRO_SOURCE_URL = (
    "https://www2.census.gov/programs-surveys/demo/tables/p60/287/"
    "SPM-pov-threshold-2024.xlsx"
)

PAPER_URL = "https://spm-threshold-paper.vercel.app"


def generate_all_thresholds():
    """Published thresholds, the packaged nowcast, then price forecasts.

    Nowcast years use the consumption-based estimate (see
    ``spm_calculator.nowcast``); remaining future years fall back to
    the package's CPI-projection path.
    """
    all_thresholds = dict(HISTORICAL_THRESHOLDS)
    nowcast_years = set(get_nowcast_years())
    for year in range(LATEST_PUBLISHED_YEAR + 1, 2031):
        if year in nowcast_years:
            all_thresholds[str(year)] = dict(
                nowcast_with_metadata(year)["values"]
            )
        else:
            all_thresholds[str(year)] = forecast_thresholds(year)
    return all_thresholds


def _load_workbook_from_bytes(content: bytes):
    from io import BytesIO

    return openpyxl.load_workbook(BytesIO(content), data_only=True)


def generate_metro_data() -> dict:
    try:
        response = requests.get(METRO_SOURCE_URL, timeout=60)
        response.raise_for_status()
    except Exception as error:
        fallback = (
            Path(__file__).resolve().parent.parent
            / "web"
            / "public"
            / "data"
            / "metro_geoadj.json"
        )
        print(
            f"Census metro workbook unavailable ({error}); reusing "
            f"committed {fallback.name}"
        )
        return json.loads(fallback.read_text())
    workbook = _load_workbook_from_bytes(response.content)
    sheet = workbook["Thresholds 2024"]

    # The Census metro workbook predates the 2026-07-17 correction, so
    # its per-tenure metro thresholds embed the pre-correction national
    # base. Ratios must use the same-vintage denominator: dividing by
    # the corrected national would silently mix vintages inside each
    # adjustment factor. Composed thresholds therefore equal the
    # workbook rescaled onto the corrected base until Census re-releases
    # the metro workbook (see tests/test_calculator.py).
    national_thresholds = get_thresholds(
        2024, series="census-published-pre-correction"
    )
    metro_areas = {}

    for row in sheet.iter_rows(min_row=3, values_only=True):
        code = row[0]
        if code is None or row[2] is None:
            continue

        rent_index = float(row[2])
        reference_thresholds = {
            "owner_with_mortgage": int(row[3]),
            "owner_without_mortgage": int(row[4]),
            "renter": int(row[5]),
        }
        adjustments = {
            tenure: reference_thresholds[tenure] / national_thresholds[tenure]
            for tenure in national_thresholds
        }

        metro_areas[str(int(code)).zfill(4 if int(code) < 10000 else 5)] = {
            "name": row[1],
            "rentIndex": rent_index,
            "adjustments": adjustments,
            "referenceThresholds": reference_thresholds,
        }

    return {
        "year": 2024,
        "source": "Census Bureau SPM Thresholds by Metro Area 2024",
        "sourceUrl": METRO_SOURCE_URL,
        "nationalThresholds": national_thresholds,
        "housingShares": TENURE_HOUSING_SHARES,
        "metroAreas": metro_areas,
    }


def generate_data():
    repo_root = Path(__file__).resolve().parent.parent
    package_data_dir = repo_root / "spm_calculator" / "data"
    web_data_dir = repo_root / "web" / "public" / "data"
    package_data_dir.mkdir(parents=True, exist_ok=True)
    web_data_dir.mkdir(parents=True, exist_ok=True)

    metro_data = generate_metro_data()
    all_thresholds = generate_all_thresholds()

    with open(package_data_dir / "metro_geoadj_2024.json", "w") as f:
        json.dump(metro_data, f, indent=2)

    with open(web_data_dir / "metro_geoadj.json", "w") as f:
        json.dump(metro_data, f, indent=2)

    metro_year = int(metro_data["year"])
    config = {
        "packageVersion": _read_package_version(),
        "baseThresholds": all_thresholds,
        "methodology": {
            "referenceRawScale": REFERENCE_RAW_SCALE,
            "housingShares": TENURE_HOUSING_SHARES,
            "equivalenceScale": {
                "singleAdultFirstChild": 0.8,
                "additionalChild": 0.5,
                "economiesOfScale": 0.7,
                "twoAdultNoChild": 1.41,
                "referenceFamilyRaw": REFERENCE_RAW_SCALE,
            },
        },
        "forecast": {
            "latestPublishedYear": LATEST_PUBLISHED_YEAR,
            "cpiProjections": {str(k): v for k, v in CPI_PROJECTIONS.items()},
        },
        # Consumption-based nowcasts for years whose CE data and CPI are
        # published but whose BLS thresholds are not. Carries the full
        # packaged document (values, components, method, caveats, label)
        # so UI disclaimers derive from the artifact.
        "nowcast": {
            str(year): nowcast_with_metadata(year)
            for year in get_nowcast_years()
        },
        "paperUrl": PAPER_URL,
        # Web mirrors the Python rule: historical years (< earliest
        # bundled) must raise rather than silently apply current-year
        # rent indices to earlier base thresholds; forecast years
        # (> latest bundled) warn-and-pin to the most recent vintage.
        "metroData": {
            "availableYears": [metro_year],
            "earliestYear": metro_year,
            "latestYear": metro_year,
        },
    }

    # `base_thresholds.json` used to be written here alongside
    # `spm_config.json`, but the web app only reads the config file
    # (which already carries `baseThresholds`). Dropping the duplicate
    # file avoids a drift risk if someone updates one without the other.
    with open(web_data_dir / "spm_config.json", "w") as f:
        json.dump(config, f, indent=2)

    print(f"Wrote {package_data_dir / 'metro_geoadj_2024.json'}")
    print(f"Wrote {web_data_dir / 'metro_geoadj.json'}")
    print(f"Wrote {web_data_dir / 'spm_config.json'}")


if __name__ == "__main__":
    generate_data()
